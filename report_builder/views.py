from django.contrib.contenttypes.models import ContentType
from django.conf import settings
from django.core.files.base import ContentFile
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import permission_required
from django.db.models import Q
from django.db.models.fields.related import ReverseManyRelatedObjectsDescriptor
from django.forms.models import inlineformset_factory
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import (
    render_to_response,
    redirect,
    get_object_or_404,
    render)
from django.template import RequestContext
from .models import Report, TabbedReport, DisplayField, FilterField, Format
from .utils import (
    javascript_date_format,
    duplicate,
)
from django.utils.decorators import method_decorator
from django.views.generic.edit import CreateView, UpdateView
from django.views.generic import TemplateView, View
from django import forms

from report_utils.model_introspection import get_relation_fields_from_model
from report_utils.mixins import GetFieldsMixin, DataExportMixin

import warnings
import datetime
import time
import re
from decimal import Decimal
from numbers import Number
import copy
from dateutil import parser
import json


class ReportForm(forms.ModelForm):
    class Meta:
        model = Report
        fields = ['name', 'distinct', 'root_model']


class ReportEditForm(forms.ModelForm):
    class Meta:
        model = Report
        fields = ['name', 'distinct', 'description']
        widgets = {
            'description': forms.TextInput(
                attrs={'style': 'width:99%;', 'placeholder': 'Description'}),
        }


class DisplayFieldForm(forms.ModelForm):
    class Meta:
        model = DisplayField
        fields = ('name', 'path', 'path_verbose', 'field_verbose', 'field', 'position',
                  'width', 'total', 'sort', 'sort_reverse', 'aggregate', 'group', 'display_format')
        widgets = {
            'path': forms.HiddenInput(),
            'path_verbose': forms.TextInput(attrs={'readonly': 'readonly'}),
            'field_verbose': forms.TextInput(attrs={'readonly': 'readonly'}),
            'field': forms.HiddenInput(),
            'width': forms.TextInput(attrs={'class': 'small_input'}),
            'total': forms.CheckboxInput(attrs={'class': 'small_input'}),
            'sort': forms.TextInput(attrs={'class': 'small_input'}),
            'sort_reverse': forms.CheckboxInput(attrs={'class': 'small_input'}),
        }


class FilterFieldForm(forms.ModelForm):
    class Meta:
        model = FilterField
        fields = ('path', 'path_verbose', 'field_verbose', 'field', 'filter_type',
                  'filter_value', 'filter_value2', 'exclude', 'position')
        widgets = {
            'path': forms.HiddenInput(),
            'path_verbose': forms.TextInput(attrs={'readonly': 'readonly'}),
            'field_verbose': forms.TextInput(attrs={'readonly': 'readonly'}),
            'field': forms.HiddenInput(),
            'filter_type': forms.Select(attrs={'onchange': 'check_filter_type(event.target)'})
        }

    def __init__(self, *args, **kwargs):
        super(FilterFieldForm, self).__init__(*args, **kwargs)
        # override the filter_value field with the models native ChoiceField
        if self.instance.choices:
            self.fields['filter_value'].widget = forms.Select(choices=self.instance.choices)
        if 'DateField' in self.instance.field_verbose or 'DateTimeField' in self.instance.field_verbose:
            widget = self.fields['filter_value'].widget
            widget.attrs['class'] = 'datepicker'
            widget.attrs['data-date-format'] = javascript_date_format(settings.DATE_FORMAT)


class ReportCreateView(CreateView):
    form_class = ReportForm
    template_name = 'report_new.html'


def filter_property(filter_field, value):
    filter_type = filter_field.filter_type
    filter_value = filter_field.filter_value
    filtered = True
    # TODO: i10n
    WEEKDAY_INTS = {
        'monday': 0,
        'tuesday': 1,
        'wednesday': 2,
        'thursday': 3,
        'friday': 4,
        'saturday': 5,
        'sunday': 6,
    }
    # TODO instead of catch all, deal with all cases
    # Example is 'a' < 2 is a valid python comparison
    # But what about 2 < '1' which yeilds true! Not intuitive for humans.
    try:
        if filter_type == 'exact' and str(value) == filter_value:
            filtered = False
        if filter_type == 'iexact' and str(value).lower() == str(filter_value).lower():
            filtered = False
        if filter_type == 'contains' and filter_value in value:
            filtered = False
        if filter_type == 'icontains' and str(filter_value).lower() in str(value).lower():
            filtered = False
        if filter_type == 'in' and value in filter_value:
            filtered = False
        # convert dates and datetimes to timestamps in order to compare digits and date/times the same
        if isinstance(value, datetime.datetime) or isinstance(value, datetime.date):
            value = str(time.mktime(value.timetuple()))
            try:
                filter_value_dt = parser.parse(filter_value)
                filter_value = str(time.mktime(filter_value_dt.timetuple()))
            except ValueError:
                pass
        if filter_type == 'gt' and Decimal(value) > Decimal(filter_value):
            filtered = False
        if filter_type == 'gte' and Decimal(value) >= Decimal(filter_value):
            filtered = False
        if filter_type == 'lt' and Decimal(value) < Decimal(filter_value):
            filtered = False
        if filter_type == 'lte' and Decimal(value) <= Decimal(filter_value):
            filtered = False
        if filter_type == 'startswith' and str(value).startswith(str(filter_value)):
            filtered = False
        if filter_type == 'istartswith' and str(value).lower().startswith(str(filter_value)):
            filtered = False
        if filter_type == 'endswith' and str(value).endswith(str(filter_value)):
            filtered = False
        if filter_type == 'iendswith' and str(value).lower().endswith(str(filter_value)):
            filtered = False
        if filter_type == 'range' and value in [int(x) for x in filter_value]:
            filtered = False
        if filter_type == 'week_day' and WEEKDAY_INTS.get(str(filter_value).lower()) == value.weekday:
            filtered = False
        if filter_type == 'isnull' and value is None:
            filtered = False
        if filter_type == 'regex' and re.search(filter_value, value):
            filtered = False
        if filter_type == 'iregex' and re.search(filter_value, value, re.I):
            filtered = False
    except:
        pass

    if filter_field.exclude:
        return not filtered
    return filtered


class AjaxGetRelated(GetFieldsMixin, TemplateView):
    template_name = "report_builder/report_form_related_li.html"

    def get_context_data(self, **kwargs):
        context = super(AjaxGetRelated, self).get_context_data(**kwargs)
        request = self.request
        model_class = ContentType.objects.get(pk=request.GET['model']).model_class()
        path = request.GET['path']
        path_verbose = request.GET['path_verbose']

        new_fields, model_ct, path = self.get_related_fields(
            model_class,
            request.GET['field'],
            path,
            path_verbose,)
        context['model_ct'] = model_ct
        context['related_fields'] = new_fields
        context['path'] = path
        context['path_verbose'] = path_verbose
        return context


def fieldset_string_to_field(fieldset_dict, model):
    if isinstance(fieldset_dict['fields'], tuple):
        fieldset_dict['fields'] = list(fieldset_dict['fields'])
    i = 0
    for dict_field in fieldset_dict['fields']:
        if isinstance(dict_field, basestring):
            fieldset_dict['fields'][i] = model._meta.get_field_by_name(dict_field)[0]
        elif isinstance(dict_field, list) or isinstance(dict_field, tuple):
            dict_field[1]['recursive'] = True
            fieldset_string_to_field(dict_field[1], model)
        i += 1


def get_fieldsets(model):
    """ fieldsets are optional, they are defined in the Model.
    """
    fieldsets = getattr(model, 'report_builder_fieldsets', None)
    if fieldsets:
        for fieldset_name, fieldset_dict in model.report_builder_fieldsets:
            fieldset_string_to_field(fieldset_dict, model)
    return fieldsets


class AjaxGetFields(GetFieldsMixin, TemplateView):
    """ Get fields from a particular model """
    template_name = 'report_builder/report_form_fields_li.html'

    def get_context_data(self, **kwargs):
        context = super(AjaxGetFields, self).get_context_data(**kwargs)
        field_name = self.request.GET.get('field')
        model_class = ContentType.objects.get(pk=self.request.GET['model']).model_class()
        path = self.request.GET['path']
        path_verbose = self.request.GET.get('path_verbose')
        root_model = model_class.__name__.lower()

        field_data = self.get_fields(model_class, field_name, path, path_verbose)
        return dict(context.items() + field_data.items())


@staff_member_required
def ajax_get_choices(request):
    path_verbose = request.GET.get('path_verbose')
    label = request.GET.get('label')
    root_model = request.GET.get('root_model')
    app_label = request.GET.get('app_label')
    model_name = path_verbose or root_model
    model_name = model_name.split(':')[-1]
    model = ContentType.objects.get(model=model_name, app_label=app_label).model_class()
    choices = FilterField().get_choices(model, label)
    select_widget = forms.Select(choices=[('', '---------')] + list(choices))
    options_html = select_widget.render_options([], [0])
    return HttpResponse(options_html)


@staff_member_required
def ajax_get_formats(request):
    choices = Format.objects.values_list('pk', 'name')
    select_widget = forms.Select(choices=[('', '---------')] + list(choices))
    options_html = select_widget.render_options([], [0])
    return HttpResponse(options_html)


class AjaxPreview(DataExportMixin, TemplateView):
    """ This view is intended for a quick preview useful when debugging
    reports. It limits to 50 objects.
    """
    template_name = "report_builder/html_report.html"

    @method_decorator(staff_member_required)
    def dispatch(self, *args, **kwargs):
        return super(AjaxPreview, self).dispatch(*args, **kwargs)

    def post(self, request, *args, **kwargs):
        return self.get(self, request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(AjaxPreview, self).get_context_data(**kwargs)
        report = get_object_or_404(Report, pk=self.request.POST['report_id'])
        queryset, message = report.get_query()
        property_filters = report.filterfield_set.filter(
            Q(field_verbose__contains='[property]') | Q(field_verbose__contains='[custom')
        )
        objects_list, message = self.report_to_list(
            queryset,
            report.displayfield_set.all(),
            self.request.user,
            property_filters=property_filters,
            preview=True,)

        context['report'] = report
        context['objects_dict'] = objects_list
        context['message'] = message
        return context


class ReportUpdateView(GetFieldsMixin, UpdateView):
    """ This view handles the edit report builder
    It includes attached formsets for display and criteria fields
    """
    model = Report
    form_class = ReportEditForm
    success_url = './'

    @method_decorator(permission_required('report_builder.change_report'))
    def dispatch(self, request, *args, **kwargs):
        return super(ReportUpdateView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super(ReportUpdateView, self).get_context_data(**kwargs)
        model_class = self.object.root_model.model_class()
        model_ct = ContentType.objects.get_for_model(model_class)
        relation_fields = get_relation_fields_from_model(model_class)

        DisplayFieldFormset = inlineformset_factory(
            Report,
            DisplayField,
            extra=0,
            can_delete=True,
            form=DisplayFieldForm)

        FilterFieldFormset = inlineformset_factory(
            Report,
            FilterField,
            extra=0,
            can_delete=True,
            form=FilterFieldForm)

        if self.request.POST:
            ctx['field_list_formset'] = DisplayFieldFormset(self.request.POST, instance=self.object)
            ctx['field_filter_formset'] = FilterFieldFormset(self.request.POST, instance=self.object, prefix="fil")
        else:
            ctx['field_list_formset'] = DisplayFieldFormset(instance=self.object)
            ctx['field_filter_formset'] = FilterFieldFormset(instance=self.object, prefix="fil")

        ctx['related_fields'] = relation_fields
        ctx['fieldsets'] = get_fieldsets(model_class)
        ctx['model_ct'] = model_ct
        ctx['root_model'] = model_ct.model
        ctx['app_label'] = model_ct.app_label

        if getattr(settings, 'REPORT_BUILDER_ASYNC_REPORT', False):
            ctx['async_report'] = True

        field_context = self.get_fields(model_class)
        ctx = dict(ctx.items() + field_context.items())

        return ctx

    def form_valid(self, form):
        context = self.get_context_data()
        field_list_formset = context['field_list_formset']
        field_filter_formset = context['field_filter_formset']

        if field_list_formset.is_valid() and field_filter_formset.is_valid():
            self.object = form.save()
            field_list_formset.report = self.object
            field_list_formset.save()
            field_filter_formset.report = self.object
            field_filter_formset.save()
            self.object.check_report_display_field_positions()
            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))


def report_to_list(report, user, preview=False, queryset=None):
    """ Create list from a report with all data filtering
    preview: Return only first 50
    objects: Provide objects for list, instead of running filters
    Returns list, message in case of issues
    """
    message = ""
    model_class = report.root_model.model_class()
    if queryset is not None:
        objects = report.add_aggregates(queryset)
    else:
        try:
            objects, query_message = report.get_query()
            message += query_message
        except exceptions.ValidationError as e:
            message += "Validation Error: {0!s}. Something may be wrong with the report's filters.".format(e)
            return [], message
        except ValueError as e:
            message += "Value Error: {0!s}. Something may be wrong with the report's filters. For example it may be expecting a number but received a character.".format(e)
            return [], message

    # Display Values
    display_field_paths = []
    property_list = {}
    custom_list = {}
    display_totals = {}

    def append_display_total(display_totals, display_field, display_field_key):
        if display_field.total:
            display_totals[display_field_key] = {'val': Decimal('0.00')}

    for i, display_field in enumerate(report.displayfield_set.all()):
        model = get_model_from_path_string(model_class, display_field.path)
        if (
            user.has_perm(model._meta.app_label + '.change_' + model._meta.module_name)
            or user.has_perm(model._meta.app_label + '.view_' + model._meta.module_name)
            or not model
        ):
            # TODO: clean this up a bit
            display_field_key = display_field.path + display_field.field
            if '[property]' in display_field.field_verbose:
                property_list[i] = display_field_key
                append_display_total(display_totals, display_field, display_field_key)
            elif '[custom' in display_field.field_verbose:
                custom_list[i] = display_field_key
                append_display_total(display_totals, display_field, display_field_key)
            elif display_field.aggregate == "Avg":
                display_field_key += '__avg'
                display_field_paths += [display_field_key]
                append_display_total(display_totals, display_field, display_field_key)
            elif display_field.aggregate == "Max":
                display_field_key += '__max'
                display_field_paths += [display_field_key]
                append_display_total(display_totals, display_field, display_field_key)
            elif display_field.aggregate == "Min":
                display_field_key += '__min'
                display_field_paths += [display_field_key]
                append_display_total(display_totals, display_field, display_field_key)
            elif display_field.aggregate == "Count":
                display_field_key += '__count'
                display_field_paths += [display_field_key]
                append_display_total(display_totals, display_field, display_field_key)
            elif display_field.aggregate == "Sum":
                display_field_key += '__sum'
                display_field_paths += [display_field_key]
                append_display_total(display_totals, display_field, display_field_key)
            else:
                display_field_paths += [display_field_key]
                append_display_total(display_totals, display_field, display_field_key)
        else:
            message += "You don't have permission to " + display_field.name
    try:
        if (
            user.has_perm(report.root_model.app_label + '.change_' + report.root_model.model)
            or user.has_perm(report.root_model.app_label + '.view_' + report.root_model.model)
        ):

            def increment_total(display_field_key, display_totals, val):
                if display_field_key in display_totals:
                    # Booleans are Numbers - blah
                    if isinstance(val, Number) and not isinstance(val, bool):
                        # do decimal math for all numbers
                        display_totals[display_field_key]['val'] += Decimal(str(val))
                    else:
                        display_totals[display_field_key]['val'] += Decimal('1.00')

            # get pk for primary and m2m relations in order to retrieve objects
            # for adding properties to report rows
            display_field_paths.insert(0, 'pk')
            m2m_relations = []
            for position, property_path in property_list.iteritems():
                property_root = property_path.split('__')[0]
                root_class = report.root_model.model_class()
                property_root_class = getattr(root_class, property_root)
                if type(property_root_class) == ReverseManyRelatedObjectsDescriptor:
                    display_field_paths.insert(1, '%s__pk' % property_root)
                    m2m_relations.append(property_root)
            values_and_properties_list = []
            filtered_report_rows = []
            group = None
            for df in report.displayfield_set.all():
                if df.group:
                    group = df.path + df.field
                    break
            if group:
                filtered_report_rows = report.add_aggregates(objects.values_list(group))
            else:
                values_list = objects.values_list(*display_field_paths)

            if not group:
                for row in values_list:
                    row = list(row)
                    obj = report.root_model.model_class().objects.get(pk=row.pop(0))
                    # related_objects
                    remove_row = False
                    values_and_properties_list.append(row)
                    # filter properties (remove rows with excluded properties)
                    property_filters = report.filterfield_set.filter(
                        Q(field_verbose__contains='[property]') | Q(field_verbose__contains='[custom')
                    )
                    for property_filter in property_filters:
                        root_relation = property_filter.path.split('__')[0]
                        if root_relation in m2m_relations:
                            pk = row[0]
                            if pk is not None:
                                # a related object exists
                                m2m_obj = getattr(obj, root_relation).get(pk=pk)
                                val = reduce(getattr, [property_filter.field], m2m_obj)
                            else:
                                val = None
                        else:
                            if '[custom' in property_filter.field_verbose:
                                for relation in property_filter.path.split('__'):
                                    if hasattr(obj, root_relation):
                                        obj = getattr(obj, root_relation)
                                val = obj.get_custom_value(property_filter.field)
                            else:
                                val = reduce(getattr, (property_filter.path + property_filter.field).split('__'), obj)
                        if filter_property(property_filter, val):
                            remove_row = True
                            values_and_properties_list.pop()
                            break
                    if not remove_row:
                        # increment totals for fields
                        for i, field in enumerate(display_field_paths[1:]):
                            if field in display_totals.keys():
                                increment_total(field, display_totals, row[i])
                        for position, display_property in property_list.iteritems():
                            relations = display_property.split('__')
                            root_relation = relations[0]
                            if root_relation in m2m_relations:
                                pk = row.pop(0)
                                if pk is not None:
                                    # a related object exists
                                    m2m_obj = getattr(obj, root_relation).get(pk=pk)
                                    val = reduce(getattr, relations[1:], m2m_obj)
                                else:
                                    val = None
                            else:
                                try:  # Could error if a related field doesn't exist
                                    val = reduce(getattr, relations, obj)
                                except AttributeError:
                                    val = None
                            values_and_properties_list[-1].insert(position, val)
                            increment_total(display_property, display_totals, val)
                        for position, display_custom in custom_list.iteritems():
                            val = obj.get_custom_value(display_custom)
                            values_and_properties_list[-1].insert(position, val)
                            increment_total(display_custom, display_totals, val)
                        filtered_report_rows += [values_and_properties_list[-1]]
                    if preview and len(filtered_report_rows) == 50:
                        break
            sort_fields = report.displayfield_set.filter(sort__gt=0).order_by('-sort').\
                values_list('position', 'sort_reverse')
            for sort_field in sort_fields:
                try:
                    filtered_report_rows = sorted(
                        filtered_report_rows,
                        key=lambda x: sort_helper(x, sort_field[0] - 1),
                        reverse=sort_field[1]
                    )
                except TypeError:  # Sorry crappy way to determine if date is being sorted
                    filtered_report_rows = sorted(
                        filtered_report_rows,
                        key=lambda x: sort_helper(x, sort_field[0] - 1, date_field=True),
                        reverse=sort_field[1]
                    )
            values_and_properties_list = filtered_report_rows
        else:
            values_and_properties_list = []
            message = "Permission Denied on %s" % report.root_model.name

        # add choice list display and display field formatting
        choice_lists = {}
        display_formats = {}
        final_list = []
        for df in report.displayfield_set.all():
            if df.choices:
                df_choices = df.choices_dict
                # Insert blank and None as valid choices
                df_choices[''] = ''
                df_choices[None] = ''
                choice_lists.update({df.position: df_choices})
            if df.display_format:
                display_formats.update({df.position: df.display_format})

        for row in values_and_properties_list:
            # add display totals for grouped result sets
            # TODO: dry this up, duplicated logic in non-grouped total routine
            if group:
                # increment totals for fields
                for i, field in enumerate(display_field_paths[1:]):
                    if field in display_totals.keys():
                        increment_total(field, display_totals, row[i])
            row = list(row)
            for position, choice_list in choice_lists.iteritems():
                row[position - 1] = unicode(choice_list[row[position - 1]])
            for position, display_format in display_formats.iteritems():
                # convert value to be formatted into Decimal in order to apply
                # numeric formats
                try:
                    value = Decimal(row[position - 1])
                except:
                    value = row[position - 1]
                # Try to format the value, let it go without formatting for ValueErrors
                try:
                    row[position - 1] = display_format.string.format(value)
                except ValueError:
                    row[position - 1] = value
            final_list.append(row)
        values_and_properties_list = final_list

        if display_totals:
            display_totals_row = []

            fields_and_properties = list(display_field_paths[1:])
            for position, value in property_list.iteritems():
                fields_and_properties.insert(position, value)
            for i, field in enumerate(fields_and_properties):
                if field in display_totals.keys():
                    display_totals_row += [display_totals[field]['val']]
                else:
                    display_totals_row += ['']

            # add formatting to display totals
            for df in report.displayfield_set.all():
                if df.display_format:
                    try:
                        value = Decimal(display_totals_row[df.position - 1])
                    except:
                        value = display_totals_row[df.position - 1]
                    # Fall back to original value if format string and value
                    # aren't compatible, e.g. a numerically-oriented format
                    # string with value which is not numeric.
                    try:
                        value = df.display_format.string.format(value)
                    except ValueError:
                        pass
                    display_totals_row[df.position - 1] = value

        if display_totals:
            values_and_properties_list = (
                values_and_properties_list + [
                    ['TOTALS'] + (len(fields_and_properties) - 1) * ['']
                ] + [display_totals_row]
            )

    except exceptions.FieldError as e:
        warnings.warn('Error {0}'.format(str(e)))
        message += "Field Error. If you are using the report builder then you found a bug!"
        message += "If you made this in admin, then you probably did something wrong."
        values_and_properties_list = None

    return values_and_properties_list, message


# TODO: tabbed report to class view
@staff_member_required
def download_tabbed_xlsx(request, pk, queryset=None):
    """ Download the full report in xlsx format
    Why xlsx? Because there is no decent ods library for python and xls has limitations
    queryset: predefined queryset to bypass filters
    """
    from openpyxl.workbook import Workbook
    import re

    report = get_object_or_404(TabbedReport, pk=pk)

    wb = Workbook(encoding='utf-8')

    for tab in report.tabs.all():
        add_report_to_workbook(wb, tab, request.user, queryset=queryset)

    # Remove default sheet
    wb.remove_sheet(wb.get_active_sheet())
    filename = re.sub(r'\W+', '', report.name) + '.xlsx'
    return get_workbook_result(wb, filename)


def get_workbook_result(wb, filename):
    from six import BytesIO
    from openpyxl.writer.excel import save_virtual_workbook

    myfile = BytesIO()
    myfile.write(save_virtual_workbook(wb))
    response = HttpResponse(
        myfile.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    response['Content-Length'] = myfile.tell()
    return response


def add_report_to_workbook(wb, report, user, queryset=None):
    """ Generate a worksheet for a report
    """
    from openpyxl.cell import get_column_letter
    import re

    ws = wb.create_sheet(title=re.sub(r'\W+', '', report.name)[:30])
    auto_width_columns = {}

    for i, field in enumerate(report.displayfield_set.all()):
        cell = ws.cell(row=0, column=i)
        cell.value = field.name
        cell.style.font.bold = True
        if field.width == 0:
            # Auto width
            auto_width_columns[i] = 1
        ws.column_dimensions[get_column_letter(i + 1)].width = field.width

    objects_list, message = report_to_list(report, user, queryset=queryset)
    for row in objects_list:
        try:
            ws.append(row)
            for col in auto_width_columns.keys():
                size = len(row[col])
                if size > auto_width_columns[col]:
                    auto_width_columns[col] = size
        except ValueError as e:
            ws.append([e.message])
        except:
            ws.append(['Unknown Error'])

    for col, width in auto_width_columns.items():
        ws.column_dimensions[get_column_letter(col + 1)].width = width

    return ws


class DownloadXlsxView(DataExportMixin, View):
    @method_decorator(staff_member_required)
    def dispatch(self, *args, **kwargs):
        return super(DownloadXlsxView, self).dispatch(*args, **kwargs)

    def process_report(self, report_id, user_id, to_response):
        report = get_object_or_404(Report, pk=report_id)
        user = get_user_model().objects.get(pk=user_id)
        queryset, message = report.get_query()
        property_filters = report.filterfield_set.filter(
            Q(field_verbose__contains='[property]') | Q(field_verbose__contains='[custom')
        )
        objects_list, message = self.report_to_list(
            queryset,
            report.displayfield_set.all(),
            user,
            property_filters=property_filters,
            preview=False,)
        title = re.sub(r'\W+', '', report.name)[:30]
        header = []
        widths = []
        for field in report.displayfield_set.all():
            header.append(field.name)
            widths.append(field.width)

        if to_response:
            return self.list_to_xlsx_response(objects_list, title, header, widths)
        else:
            self.async_report_save(report, objects_list, title, header, widths)

    def async_report_save(self, report, objects_list, title, header, widths):
        xlsx_file = self.list_to_xlsx_file(objects_list, title, header, widths)
        if not title.endswith('.xlsx'):
            title += '.xlsx'
        report.report_file.save(title, ContentFile(xlsx_file.getvalue()))
        report.report_file_creation = datetime.datetime.today()
        report.save()

    def get(self, request, *args, **kwargs):
        report_id = kwargs['pk']
        if getattr(settings, 'REPORT_BUILDER_ASYNC_REPORT', False):
            from .tasks import report_builder_async_report_save
            report_task = report_builder_async_report_save.delay(report_id, request.user.pk)
            task_id = report_task.task_id
            return HttpResponse(json.dumps({'task_id': task_id}), content_type="application/json")
        else:
            return self.process_report(report_id, request.user.pk, to_response=True)


@staff_member_required
def ajax_add_star(request, pk):
    """ Star or unstar report for user
    """
    report = get_object_or_404(Report, pk=pk)
    user = request.user
    if user in report.starred.all():
        added = False
        report.starred.remove(request.user)
    else:
        added = True
        report.starred.add(request.user)
    return HttpResponse(added)


@staff_member_required
def create_copy(request, pk):
    """ Copy a report including related fields """
    report = get_object_or_404(Report, pk=pk)
    new_report = duplicate(report, changes=(
        ('name', '{0} (copy)'.format(report.name)),
        ('user_created', request.user),
        ('user_modified', request.user),
    ))
    # duplicate does not get related
    for display in report.displayfield_set.all():
        new_display = copy.copy(display)
        new_display.pk = None
        new_display.report = new_report
        new_display.save()
    for report_filter in report.filterfield_set.all():
        new_filter = copy.copy(report_filter)
        new_filter.pk = None
        new_filter.report = new_report
        new_filter.save()
    return redirect(new_report)


@staff_member_required
def export_to_report(request):
    """ Export objects (by ID and content type) to an existing or new report
    In effect this runs the report with it's display fields. It ignores
    filters and filters instead the provided ID's. It can be select
    as a global admin action.
    """
    admin_url = request.GET.get('admin_url', '/')
    ct = ContentType.objects.get_for_id(request.GET['ct'])
    ids = request.GET['ids'].split(',')
    number_objects = len(ids)
    reports = Report.objects.filter(root_model=ct).order_by('-modified')

    if 'download' in request.GET:
        report = get_object_or_404(Report, pk=request.GET['download'])
        queryset = ct.model_class().objects.filter(pk__in=ids)
        return download_xlsx(request, report.id, queryset=queryset)

    return render(request, 'report_builder/export_to_report.html', {
        'object_list': reports,
        'admin_url': admin_url,
        'number_objects': number_objects,
        'model': ct.model_class()._meta.verbose_name,
    })


@staff_member_required
def check_status(request, pk, task_id):
    """ Check if the asyncronous report is ready to download """
    from celery.result import AsyncResult
    res = AsyncResult(task_id)
    link = ''
    if res.state == 'SUCCESS':
        report = get_object_or_404(Report, pk=pk)
        link = report.report_file.url
    return HttpResponse(json.dumps({'state': res.state, 'link': link}), content_type="application/json")
